#!/usr/bin/env python3
import pandas as pd
import openpyxl
import numpy as np
from datetime import datetime
import re

def analyze_excel_file(file_path):
    """Analyze Excel file structure and data formats"""
    
    print(f"Analyzing file: {file_path}")
    print("=" * 80)
    
    # Load the workbook with openpyxl for sheet information
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print(f"\nNumber of sheets: {len(wb.sheetnames)}")
    print(f"Sheet names: {', '.join(wb.sheetnames)}")
    
    # Analyze each sheet
    for sheet_name in wb.sheetnames:
        print(f"\n\n{'='*80}")
        print(f"ANALYZING SHEET: {sheet_name}")
        print(f"{'='*80}")
        
        # Read the sheet with pandas
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)
        
        print(f"\nShape: {df.shape[0]} rows × {df.shape[1]} columns")
        print(f"\nColumn Names:")
        for i, col in enumerate(df.columns, 1):
            # Check if column name starts with a number
            starts_with_number = bool(re.match(r'^\d', str(col)))
            print(f"  {i}. '{col}' {' [STARTS WITH NUMBER]' if starts_with_number else ''}")
        
        print(f"\n\nDetailed Column Analysis:")
        print("-" * 80)
        
        for col in df.columns:
            print(f"\nColumn: '{col}'")
            print(f"  Non-null count: {df[col].count()} / {len(df)}")
            print(f"  Null count: {df[col].isnull().sum()}")
            print(f"  Data type (pandas): {df[col].dtype}")
            
            # Get unique non-null values (up to 10 examples)
            non_null_values = df[col].dropna()
            if len(non_null_values) > 0:
                unique_values = non_null_values.unique()
                print(f"  Unique values: {len(unique_values)}")
                
                # Show sample values
                sample_size = min(5, len(unique_values))
                print(f"  Sample values (first {sample_size}):")
                for val in unique_values[:sample_size]:
                    print(f"    - '{val}' (type: {type(val).__name__})")
                
                # Analyze formats for string columns
                if df[col].dtype == 'object':
                    # Check for special patterns
                    patterns = {
                        'Date-like': r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',
                        'Number with commas': r'[\d,]+\.?\d*',
                        'Number in quotes': r'^"[\d,.]+"$',
                        'Currency': r'\$[\d,]+\.?\d*',
                        'Percentage': r'\d+\.?\d*%',
                        'Special chars': r'[^\w\s\-\.]'
                    }
                    
                    for pattern_name, pattern in patterns.items():
                        matches = non_null_values.astype(str).str.contains(pattern, na=False).sum()
                        if matches > 0:
                            print(f"    - Contains {pattern_name}: {matches} values")
                            # Show examples
                            examples = non_null_values[non_null_values.astype(str).str.contains(pattern, na=False)].head(3)
                            for ex in examples:
                                print(f"      Example: '{ex}'")
                
                # Check for numeric values stored as strings
                if df[col].dtype == 'object':
                    try:
                        # Try to convert to numeric
                        numeric_convertible = pd.to_numeric(non_null_values.astype(str).str.replace(',', '').str.replace('"', ''), errors='coerce')
                        convertible_count = numeric_convertible.notna().sum()
                        if convertible_count > 0:
                            print(f"    - Numeric values stored as text: {convertible_count}/{len(non_null_values)}")
                    except:
                        pass
            
            # Check for leading/trailing spaces
            if df[col].dtype == 'object' and len(non_null_values) > 0:
                with_spaces = non_null_values.astype(str).str.strip() != non_null_values.astype(str)
                if with_spaces.any():
                    print(f"    - Values with leading/trailing spaces: {with_spaces.sum()}")
        
        # Additional analysis
        print(f"\n\nAdditional Analysis:")
        print("-" * 80)
        
        # Check for completely empty columns
        empty_cols = [col for col in df.columns if df[col].isnull().all()]
        if empty_cols:
            print(f"Empty columns: {', '.join(empty_cols)}")
        
        # Check for duplicate column names
        duplicate_cols = [col for col in df.columns if list(df.columns).count(col) > 1]
        if duplicate_cols:
            print(f"Duplicate column names: {', '.join(set(duplicate_cols))}")
        
        # Show first few rows as preview
        print(f"\n\nData Preview (first 3 rows):")
        print("-" * 80)
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        pd.set_option('display.max_colwidth', 50)
        print(df.head(3))
        
    wb.close()

if __name__ == "__main__":
    file_path = "/mnt/c/Users/sston/Dropbox/-- Greenway/sample/greenway_2025-03-31_vc_daily_remittance_report - loan only.xlsx"
    analyze_excel_file(file_path)